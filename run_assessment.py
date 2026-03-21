import argparse
from openpyxl import load_workbook
from config import EXCEL_PATH, EXCEL_SHEET
from utils import flatten_assessment


def main():
    parser = argparse.ArgumentParser(description="Run a policy commitment assessment.")
    parser.add_argument("--commitment-id", "-c", required=True, help="Commitment ID to assess (e.g. CP.1)")
    parser.add_argument("--pdf-source", "-p", required=True, help="Path to the policy PDF file")
    parser.add_argument("--approach", "-a", default="two_step_simple", choices=["single_step", "two_step_simple", "two_step"], help="Assessment approach (default: two_step_simple)")
    parser.add_argument("--model", "-m", default="gpt-4o", help="Model name to use (default: gpt-4o)")
    parser.add_argument("--policy-debug", action="store_true", help="Print policy pages for debugging")
    parser.add_argument("--input-debug", action="store_true", help="Print other inputs for debugging")
    parser.add_argument("--interactive", action="store_true", help="Manually select and edit commitment references (not applicable for single_step)")
    parser.add_argument("--to-excel", action="store_true", help="Export two_step results to Excel")
    args = parser.parse_args()

    COMMITMENT_ID = args.commitment_id
    PDF_SOURCE = args.pdf_source
    APPROACH = args.approach
    MODEL_NAME = args.model
    POLICY_DEBUG = args.policy_debug
    INPUT_DEBUG = args.input_debug
    INTERACTIVE_MODE = args.interactive
    TO_EXCEL = args.to_excel

    print(f"\n{'='*80}")
    print(f"Running assessment with approach: {APPROACH.upper()}")
    print(f"Commitment: {COMMITMENT_ID}")
    print(f"Source: {PDF_SOURCE}")
    print(f"Model: {MODEL_NAME}")
    print(f"{'='*80}\n")

    if APPROACH == "single_step":
        from single_step import run_single_step_analysis
        result = run_single_step_analysis(
            commitment_id=COMMITMENT_ID,
            pdf_source=PDF_SOURCE,
            model_name=MODEL_NAME,
            policy_debug=POLICY_DEBUG,
            input_debug=INPUT_DEBUG,
            interactive_mode=INTERACTIVE_MODE
        )

    elif APPROACH == "two_step":
        from two_step import run_two_step_analysis
        result = run_two_step_analysis(
            commitment_id=COMMITMENT_ID,
            pdf_source=PDF_SOURCE,
            model_name=MODEL_NAME,
            policy_debug=POLICY_DEBUG,
            input_debug=INPUT_DEBUG,
            interactive_mode=INTERACTIVE_MODE
        )
        if TO_EXCEL:
            flattened = flatten_assessment(result, COMMITMENT_ID)
            target_sheet = EXCEL_SHEET or "Sheet1"

            try:
                wb = load_workbook(EXCEL_PATH)
            except FileNotFoundError:
                raise FileNotFoundError(f"Excel template not found at {EXCEL_PATH}")

            if target_sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{target_sheet}' not found in {EXCEL_PATH}")

            ws = wb[target_sheet]

            # Read header row to get column mapping
            headers = {}
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value:
                    headers[cell.value] = col_idx

            if "commitment_id" not in headers:
                raise KeyError("Target sheet must contain a 'commitment_id' column")

            # Find existing row with matching commitment_id
            commitment_col = headers["commitment_id"]
            target_row = None
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row_idx, commitment_col).value == COMMITMENT_ID:
                    target_row = row_idx
                    break

            # If no match, append new row
            if target_row is None:
                target_row = ws.max_row + 1
                row_action = "appended"
            else:
                row_action = "updated"

            # Write data to cells
            for col_name, value in flattened.items():
                if col_name in headers:
                    col_idx = headers[col_name]
                    ws.cell(target_row, col_idx, value)

            wb.save(EXCEL_PATH)
            print(f"Two-step Excel export {row_action} in {EXCEL_PATH} (sheet '{target_sheet}')")

    elif APPROACH == "two_step_simple":
        from two_step_simple import run_two_step_simple_analysis
        result = run_two_step_simple_analysis(
            commitment_id=COMMITMENT_ID,
            pdf_source=PDF_SOURCE,
            model_name=MODEL_NAME,
            policy_debug=POLICY_DEBUG,
            input_debug=INPUT_DEBUG,
            interactive_mode=INTERACTIVE_MODE
        )


if __name__ == "__main__":
    main()

