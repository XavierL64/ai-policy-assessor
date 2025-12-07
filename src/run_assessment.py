# Assessment parameters
COMMITMENT_ID = "CP.1"
PDF_SOURCE = "policies/ABN/Exclusion list (Mar 2021).pdf"

# Run config
APPROACH = "two_step" # Choose from "single_step", "two_step_simple", "two_step"
MODEL_NAME = "gpt-4.1"
POLICY_DEBUG = False  # Set to True to print policy pages
INPUT_DEBUG = False   # Set to True to print other inputs
INTERACTIVE_MODE = False  # Set to True to manually select and edit commitment references (not applicable for single_step)
TO_EXCEL = True  # Set to True to export two_step results to Excel

import pandas as pd
from openpyxl import load_workbook
from config import EXCEL_PATH, EXCEL_SHEET
from utils import flatten_assessment

if __name__ == "__main__":
    print(f"\n{'='*80}")
    print(f"Running assessment with approach: {APPROACH.upper()}")
    print(f"Commitment: {COMMITMENT_ID}")
    print(f"Source: {PDF_SOURCE}")
    print(f"Model: {MODEL_NAME}")
    print(f"{'='*80}\n")

    if APPROACH == "single_step":
        from single_step import run_single_step_analysis
        result = run_single_step_analysis(
            commitment_id=COMMITMENT_ID,
            pdf_source=PDF_SOURCE,
            model_name=MODEL_NAME,
            policy_debug=POLICY_DEBUG,
            input_debug=INPUT_DEBUG,
            interactive_mode=INTERACTIVE_MODE
        )

    elif APPROACH == "two_step":
        from two_step import run_two_step_analysis
        result = run_two_step_analysis(
            commitment_id=COMMITMENT_ID,
            pdf_source=PDF_SOURCE,
            model_name=MODEL_NAME,
            policy_debug=POLICY_DEBUG,
            input_debug=INPUT_DEBUG,
            interactive_mode=INTERACTIVE_MODE
        )
        if TO_EXCEL:
            flattened = flatten_assessment(result, COMMITMENT_ID)
            target_sheet = EXCEL_SHEET or "Sheet1"

            try:
                wb = load_workbook(EXCEL_PATH)
            except FileNotFoundError:
                raise FileNotFoundError(f"Excel template not found at {EXCEL_PATH}")

            if target_sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{target_sheet}' not found in {EXCEL_PATH}")

            ws = wb[target_sheet]

            # Read header row to get column mapping
            headers = {}
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value:
                    headers[cell.value] = col_idx

            if "commitment_id" not in headers:
                raise KeyError("Target sheet must contain a 'commitment_id' column")

            # Find existing row with matching commitment_id
            commitment_col = headers["commitment_id"]
            target_row = None
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row_idx, commitment_col).value == COMMITMENT_ID:
                    target_row = row_idx
                    break

            # If no match, append new row
            if target_row is None:
                target_row = ws.max_row + 1
                row_action = "appended"
            else:
                row_action = "updated"

            # Write data to cells
            for col_name, value in flattened.items():
                if col_name in headers:
                    col_idx = headers[col_name]
                    ws.cell(target_row, col_idx, value)

            wb.save(EXCEL_PATH)
            print(f"Two-step Excel export {row_action} in {EXCEL_PATH} (sheet '{target_sheet}')")

    elif APPROACH == "two_step_simple":
        from two_step_simple import run_two_step_simple_analysis
        result = run_two_step_simple_analysis(
            commitment_id=COMMITMENT_ID,
            pdf_source=PDF_SOURCE,
            model_name=MODEL_NAME,
            policy_debug=POLICY_DEBUG,
            input_debug=INPUT_DEBUG,
            interactive_mode=INTERACTIVE_MODE
        )

    else:
        raise ValueError(f"Unknown approach: {APPROACH}. Must be 'single_step', 'two_step_simple', or 'two_step'")
